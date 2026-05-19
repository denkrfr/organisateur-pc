"""Extraction de texte pour les documents (PDF / DOCX / XLSX).

Utilise des librairies pure-Python pour rester local et leger :
  - pypdf       : PDF
  - python-docx : Word .docx
  - openpyxl    : Excel .xlsx

Si une lib n'est pas dispo (build minimal), les fonctions renvoient None
au lieu de planter. Le pipeline degrade gracefully.
"""

from __future__ import annotations
import hashlib
import re
from pathlib import Path
from typing import Optional


DOC_EXTS = {".pdf", ".docx", ".xlsx"}
VIDEO_EXTS = {".mp4", ".mov", ".mkv", ".avi", ".webm", ".m4v", ".3gp", ".wmv", ".flv"}


def kind_of(path: Path) -> str:
    """Retourne 'image', 'pdf', 'docx', 'xlsx', 'video' ou 'other'."""
    ext = path.suffix.lower()
    if ext in {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp", ".heic", ".heif", ".tiff", ".tif"}:
        return "image"
    if ext == ".pdf":
        return "pdf"
    if ext == ".docx":
        return "docx"
    if ext == ".xlsx":
        return "xlsx"
    if ext in VIDEO_EXTS:
        return "video"
    return "other"


# ---------------------------------------------------------------------------
# Extraction texte
# ---------------------------------------------------------------------------
def extract_text_pdf(path: Path, max_chars: int = 50_000) -> Optional[str]:
    """Extrait le texte d'un PDF. None si lib absente ou fichier illisible."""
    try:
        from pypdf import PdfReader  # type: ignore[import-not-found]
    except ImportError:
        return None
    try:
        reader = PdfReader(str(path))
        chunks: list[str] = []
        for page in reader.pages:
            try:
                chunks.append(page.extract_text() or "")
            except Exception:  # noqa: BLE001 — page foireuse, on continue
                continue
            if sum(len(c) for c in chunks) >= max_chars:
                break
        return "\n".join(chunks)[:max_chars]
    except Exception:  # noqa: BLE001 — PDF chiffre / corrompu
        return None


def extract_text_docx(path: Path, max_chars: int = 50_000) -> Optional[str]:
    try:
        from docx import Document  # type: ignore[import-not-found]
    except ImportError:
        return None
    try:
        doc = Document(str(path))
        chunks: list[str] = []
        for p in doc.paragraphs:
            chunks.append(p.text)
            if sum(len(c) for c in chunks) >= max_chars:
                break
        return "\n".join(chunks)[:max_chars]
    except Exception:  # noqa: BLE001
        return None


def extract_text_xlsx(path: Path, max_chars: int = 50_000) -> Optional[str]:
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except ImportError:
        return None
    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
        chunks: list[str] = []
        total = 0
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            chunks.append(f"[Feuille: {sheet}]")
            total += len(chunks[-1])
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None and str(c).strip()]
                if not cells:
                    continue
                line = " | ".join(cells)
                chunks.append(line)
                total += len(line)
                if total >= max_chars:
                    break
            if total >= max_chars:
                break
        wb.close()
        return "\n".join(chunks)[:max_chars]
    except Exception:  # noqa: BLE001
        return None


def extract_text(path: Path, max_chars: int = 50_000) -> Optional[str]:
    """Dispatch par type."""
    k = kind_of(path)
    if k == "pdf":
        return extract_text_pdf(path, max_chars)
    if k == "docx":
        return extract_text_docx(path, max_chars)
    if k == "xlsx":
        return extract_text_xlsx(path, max_chars)
    return None


# ---------------------------------------------------------------------------
# Fingerprint texte pour quasi-doublons docs
# ---------------------------------------------------------------------------
_WS_RE = re.compile(r"\s+")


def text_fingerprint(text: str) -> Optional[str]:
    """Normalise le texte (lowercase, whitespace collapse) et SHA256.

    Deux docs avec le meme contenu mais des metadonnees differentes
    (export PDF different, version Office differente) auront le meme
    fingerprint. C'est plus strict que du fuzzy mais c'est honnete et
    rapide. Pour V2 on pourra ajouter du simhash.
    """
    if not text:
        return None
    normalized = _WS_RE.sub(" ", text.lower()).strip()
    if len(normalized) < 50:
        return None  # trop court pour etre fiable
    return hashlib.sha256(normalized.encode("utf-8", errors="replace")).hexdigest()
